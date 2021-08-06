from bot.checker import check_one
from bot import base
from bot.loger import get_logger
from openpyxl import load_workbook
from datetime import datetime

logger = get_logger(__name__)


class Prostavka:
    def __init__(self, name: str, costs: list, articul: str, krep: str, comment: str, photo: str):
        self.name = name
        self.costs = costs
        self.articul = articul
        self.krep = krep
        self.comment = comment
        self.photo = photo
        self.cost = 0

    def set_height(self, height):
        self.height = height

    def set_cost(self, cost: int):
        self.cost = cost


class Client:
    mark = ''
    model = None
    generation = None
    modification = None
    prostavki = {}
    cost = 0

    def write_itog(self):
        book = load_workbook(r"C:\Users\mv149\PycharmProjects\eazybot\Карточка клиента.xlsx")
        cards = book.get_sheet_by_name(book.get_sheet_names()[0])
        now = datetime.now().strftime("%d-%m-%Y")
        cards.append([cards.max_row - 2, now, self.generation, self.mark, self.model, self.modification,
                      self.prostavki.get('Передние проставки', ['Нет'])[0],
                      self.prostavki.get('Передние проставки', ['Нет'])[1],
                      self.prostavki.get('Задние проставки', ['Нет'])[0],
                      self.prostavki.get('Задние проставки', ['Нет'])[1],
                      self.prostavki.get('Удлинители', ['Нет'])[0],
                      self.prostavki.get('Удлинители', ['Нет'])[1],
                      self.prostavki.get('Передние проставки', ['Нет'])[2],
                      self.prostavki.get('Задние проставки', ['Нет'])[2],
                      self.prostavki.get('Удлинители', ['Нет'])[2],
                      0,
                      self.cost
                      ])
        book.save('Карточка клиента.xlsx')


def dialog():
    car = yield "Здравствуйте, для заказа напишите марку и модель вашей машины в виде \"марка модель\" (например Volvo C30, тойота камри)"
    cars = check_one(car)
    logger.info(car)
    logger.info(cars)
    s = 'Выберите подходящий вариант:\n'
    for cnt, i in enumerate(cars):
        s = s + f'{cnt + 1}) ' + i[0]
    ans = yield s
    car = cars[int(ans) - 1][0][:-1]
    client = Client()
    client.model = car
    for i in client.model.split()[:-1]:
        client.mark += i
    yield from generations(client, car)


def generations(client, car):
    gen = []
    logger.info(car)
    for i in list(base.iter_cols(min_col=2, max_col=2))[0]:
        val = base.cell(i.row, i.col_idx + 1).value
        if i.value == car and val not in gen:
            gen.append(val)
    logger.info(gen)
    if len(gen) > 1:
        s = ''
        for cnt, i in enumerate(gen):
            s = s + str(cnt + 1) + ')' + f' {i}\n'
        ans = yield "Выберите поколение вашего автомобиля(напишите его номер):\n" + s
        ans = int(ans)
        ans = gen[ans - 1]
        client.generation = ans
        yield from modif(client, ans)
    else:
        client.generation = gen[0]
        yield from modif(client, gen[0])


def modif(client, gen):
    mod = []
    for i in list(base.iter_cols(min_col=3, max_col=3))[0]:
        val = base.cell(i.row, i.col_idx + 1).value
        if i.value == gen and val not in mod:
            mod.append(val)
    if len(mod) > 1:
        s = ''
        for cnt, i in enumerate(mod):
            s = s + str(cnt + 1) + ')' + f' {i}\n'
        ans = yield "Выберите модификацию вашего автомобиля(напишите его номер):\n" + s
        ans = int(ans)
        ans = mod[ans - 1]
        client.modification = ans
        yield from choose_direction(client, ans)
    else:
        client.modification = mod[0]
        yield from choose_direction(client, mod[0])


def choose_direction(client, mod):
    dir = []
    prostavki = []
    for i in list(base.iter_cols(min_col=4, max_col=4))[0]:
        name = base.cell(i.row, i.col_idx + 1).value
        if i.value == mod and name not in dir:
            dir.append(name)
            prostavki.append(Prostavka(name,
                                       [int(base.cell(i.row, i.col_idx + 4).value or 0),
                                        int(base.cell(i.row, i.col_idx + 5).value or 0),
                                        int(base.cell(i.row, i.col_idx + 6).value or 0),
                                        int(base.cell(i.row, i.col_idx + 7).value or 0)],
                                       base.cell(i.row, i.col_idx + 2).value, base.cell(i.row, i.col_idx + 3),
                                       base.cell(i.row, i.col_idx + 8), base.cell(i.row, i.col_idx + 9)
                                       ))
    if len(dir) == 2:
        s = ''
        dir.append("Передние и задние")
        for cnt, i in enumerate(dir):
            s = s + str(cnt + 1) + ')' + f' {i}\n'
        ans = yield "Выберите тип проставок вашего автомобиля(напишите его номер):\n" + s
        ans = int(ans) - 1
        yield from get_height(client, prostavki, dir[ans])
    elif len(dir) == 3:
        s = ''
        dir.append("Передние и задние")
        dir.append("Полный комплект")
        for cnt, i in enumerate(dir):
            s = s + str(cnt + 1) + ')' + f' {i}\n'
        ans = yield "Выберите тип проставок вашего автомобиля(напишите его номер):\n" + s
        ans = int(ans) - 1
        yield from get_height(client, prostavki, dir[ans])
    else:
        yield from get_height(client, prostavki, dir[0])


def get_height(client, prostavki, dir):
    heights = []
    costs = []
    itog = []
    if dir == 'Передние проставки' or dir == 'Задние проставки' or dir == 'Удлинители':
        for i in prostavki:
            if i.name == dir:
                itog.append(i)
                for cnt, j in enumerate(i.costs):
                    if j:
                        heights.append(f'{cnt + 2}0mm')
                        costs.append(j)
    elif dir == 'Передние и задние':
        for i in prostavki:
            if i.name == 'Передние проставки' or i.name == 'Задние проставки':
                itog.append(i)
                for cnt, j in enumerate(i.costs):
                    if j:
                        try:
                            costs[cnt] += j
                        except:
                            costs.append(j)
                            heights.append(f'{cnt + 2}0mm')
    else:
        for i in prostavki:
            itog.append(i)
            for cnt, j in enumerate(i.costs):
                if j:
                    try:
                        costs[cnt] += j
                    except:
                        costs.append(j)
                        heights.append(f'{cnt + 2}0mm')
    s = ''
    for cnt, h in enumerate(heights):
        s = s + str(cnt + 1) + ')' + f' {h}\n'
    ans = yield 'Выберите высоту проставки:\n' + s
    ans = int(ans) - 1
    h = heights[ans]
    for i in itog:
        i.set_height(h)
        i.set_cost(i.costs[ans])
        client.prostavki.update({i.name: (i.height, i.articul, i.cost)})
    cost = costs[ans]
    client.cost = costs[ans]
    client.write_itog()
    yield from get_itog(cost)


def get_itog(cost):
    yield f"Итоговая цена {cost}"
